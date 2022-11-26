# -*- encoding: utf-8 -*-

from openpyxl import load_workbook


class SpecAbility:
    def __init__(self):
        self.device_type_list = {}
        self.attr_key_list = {}
        self.read_book = None
        self.table = None
        self.spec_ab_path = 'specification_info.xlsx'

        self.init()

    def init(self):
        self.read_book = load_workbook(self.spec_ab_path)
        self.table = self.read_book.get_sheet_by_name(self.read_book.get_sheet_names()[0])
        nrows = self.table.max_row  # 获取table工作表总行数
        ncols = self.table.max_column  # 获取table工作表总列数
        print "excel max_row:", nrows, "max_column:", ncols

        for j in range(1, ncols + 1):
            cell_value = self.table.cell(1, j).value
            # print cell_value
            self.attr_key_list[cell_value] = j

        type_col_index = self.attr_key_list[u'款型*']
        # print type_col_index
        for i in range(2, nrows + 1):
            cell_value = self.table.cell(i, type_col_index).value
            # print cell_value
            self.device_type_list[cell_value] = i

    def get_device_list(self):
        return self.device_type_list

    def get_device_index(self, device_type):
        if device_type in self.device_type_list:
            return self.device_type_list[device_type]
        else:
            print "==== Not Found ====", device_type
            return 0

    def get_attr_list(self):
        return self.attr_key_list

    def get_attr_index(self, attr_key):
        if attr_key in self.attr_key_list:
            return self.attr_key_list[attr_key]
        else:
            print "==== Not Found ====", attr_key
            return 0

    def write_cell(self, device_type, attr_key, value):
        type_row_index = self.get_device_index(device_type)
        type_col_index = self.get_attr_index(attr_key)
        if type_row_index == 0 and type_col_index == 0:
            print "index can't 0: ", device_type, type_row_index, type_col_index, attr_key, value
            return
        # print "write to excel:", device_type, type_row_index, type_col_index, attr_key, value
        self.table.cell(type_row_index, type_col_index).value = value

    def close(self):
        print "write done to: ", self.spec_ab_path
        self.read_book.save(self.spec_ab_path)
